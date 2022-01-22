import openpyxl

elements = []

excel_file = openpyxl.load_workbook('decmo.xlsx')

management_sheet = excel_file['january']

currently_active_sheet = excel_file.active

for i in range (2,94):
    elements = []
    for x in range(2, management_sheet.max_column+1):

        u = management_sheet.cell(row=i, column=x).value

        if u!= None:
            elements.append(u)
        else:
            break

    if len(elements)>0:
        print(elements)


