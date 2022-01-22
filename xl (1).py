import xlsxwriter
from tkinter import *

from openpyxl import load_workbook

myFileName = 'decmo.xlsx'
# load the workbook, and put the sheet into a variable
wb = load_workbook(filename=myFileName)

newRowLocation = 1

global index
index = 1


def functionate():
    v = e.get()
    vv = f.get()
    d = date.get()
    m = month.get()
    find = d + '##'
    print(find)
    x = 0
    y = 0
    c = 0
    ws = wb['january']
    if m == '1':
        ws = wb['january']
    if m == '2':
        ws = wb['february']
    if m == '3':
        ws = wb['march']
    if m == '4':
        ws = wb['april']
    if m == '5':
        ws = wb['may']
    if m == '6':
        ws = wb['june']
    if m == '7':
        ws = wb['july']
    if m == '8':
        ws = wb['august']
    if m == '9':
        ws = wb['september']
    if m == '10':
        ws = wb['october']
    if m == '11':
        ws = wb['november']
    if m == '12':
        ws = wb['december']

    global index

    for row in ws.iter_rows(min_row=1, min_col=1, max_row=30, max_col=2):
        for cell in row:
            if cell.value == find:
                print('found')
                x = cell.column
                y = cell.row
                z = cell.coordinate
                income_obj = ws.cell(row=y + 1, column=x + 1)
                cost_obj = ws.cell(row=y + 1, column=x + 2)
                c = income_obj.value
                cost=cost_obj.value

    z = ws.max_column

    ws.cell(column=c + x + 3, row=y + 1, value=v)
    ws.cell(column=c + x + 4, row=y + 1, value=vv)
    total=float(cost)+float(vv)
    ws.cell(column=x + 1, row=y + 1, value=c + 2)
    ws.cell(column=x + 2, row=y+1, value=total)

    print(c)
    wb.save(filename=myFileName)
    wb.close()


root = Tk()
root.geometry('500x500')
e = Entry(root)
b = Button(root, text='Click', command=functionate)
e.pack()
b.pack()
e.place(x=0, y=0)
b.place(x=50, y=50)

f = Entry(root)
f.pack()
date = Entry(root)
date.pack()
date.place(x=100, y=100)
month = Entry(root)
month.pack()
month.place(x=100, y=150)

print(4)

root.mainloop()
